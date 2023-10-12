#this is a simple gui to understand how PYQT5 works
import PyQt5.QtWidgets as QtWidgets
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QDialog, QApplication, QPushButton, QLabel, QGridLayout, QCheckBox, QScrollArea, QListWidget, QListWidgetItem, QTableWidget, QTableWidgetItem, QTextEdit, QMessageBox
from PyQt5 import uic
import sys
import os
import socket
import platform
import subprocess
import shutil
from PyQt5.QtGui import QDesktopServices
from PyQt5.QtCore import QUrl
import openpyxl
from openpyxl.utils import cell

# Run the 'brew --version' command and capture the output
result = subprocess.run(['brew', '--version'], capture_output=True, text=True)

# Check if Homebrew is installed
if result.returncode == 0:
    print("Homebrew is installed.")
else:
    subprocess.check_call(['/bin/bash', '-c', '$(curl -fsSL https://raw.githubusercontent.com/Homebrew/install/HEAD/install.sh)'])
    
#check if openpyxl is installed
try:
    import openpyxl
except ModuleNotFoundError:
    subprocess.check_call(['pip3', 'install', 'openpyxl'])
    import openpyxl
from openpyxl.utils import cell
from datetime import date
from openpyxl.styles import Alignment



#This is the main page that contains five buttons: Inventory(which transfers to a table view with active inventory table), Check Software(list of checkboxes that outputs what has been downloaded), Loan Agreement, Work Request Form, and Exit.
class UI(QDialog):
    data = {"Building": 'N/A',
        "Room": 'N/A',
        "Computer Name": 'N/A',
        "Tag": 'N/A',
        "S/N": 'N/A',
        "MAC": 'N/A',
        "wMac" : 'N/A',
        "First Production Date" : 'N/A',
        "Make" : 'N/A',
        "Model": 'N/A',
        "Last Service Date" : 'N/A',
        "User" : 'N/A',
        "Advisor" : 'N/A',
        "Project" : 'N/A',
        "Requistion": 'N/A',
        "Notes": 'N/A'}
        
    def __init__(self):
        super(UI, self).__init__()

        # Load the ui file
        uic.loadUi("mainscreen.ui", self)
        



        # Define Our Button Widgets
        self.software = self.findChild(QPushButton, "software")
        self.exit = self.findChild(QPushButton, "exit")
        self.inventory = self.findChild(QPushButton, "inventory")
        self.loan = self.findChild(QPushButton, "loan")
        self.work = self.findChild(QPushButton, "work_request")

        # Assign action when clicked
        self.software.clicked.connect(self.checkSoftware)
        self.exit.clicked.connect(QApplication.quit)
        self.inventory.clicked.connect(self.inInventory)
        self.loan.clicked.connect(self.LoanAgreement)

        # Show The App
        self.show()

    def finished(self):
        print("exit")

    def checkSoftware(self):
        widget.setCurrentIndex(widget.currentIndex() + 2)

    def inInventory(self):
        widget.setCurrentIndex(widget.currentIndex() + 3)
    def LoanAgreement(self):
        # set the path of the PDF file to open
        file_path = "Loan Agreement Template.pdf"

        # copy the file to a new location
        copy_path = "LoanAgreementCopy.pdf"
        shutil.copyfile(file_path, copy_path)

        # open the copied file with the default application
        QDesktopServices.openUrl(QUrl.fromLocalFile(copy_path))
        
class Tag(QDialog):
    def __init__(self):
        super(Tag, self).__init__()
        
        # Load the ui file
        uic.loadUi("tag.ui", self)

        # Define Our Button Widgets
        self.tagNumber = self.findChild(QTextEdit, "tag_number")
        self.exit = self.findChild(QPushButton, "exit")
        self.submit = self.findChild(QPushButton, "submit")

        # Assign action when clicked
        self.submit.clicked.connect(self.inventoryCheck)
        self.exit.clicked.connect(self.backToMain3)
        
        # Show The App
        self.show()

    def inventoryCheck(self):
        i = InventoryGathering()
        text = self.tagNumber.toPlainText()
        tag = str(text)
        UI.data["Tag"] = tag
        print(UI.data)
        print(tag)
        i.tagger(tag)
        inventory.update_list()
        widget.setCurrentIndex(widget.currentIndex() - 2)

    def backToMain3(self):
        widget.setCurrentIndex(widget.currentIndex() - 3)




class Inventory(QDialog):
            
    def __init__(self):
        super(Inventory, self).__init__()
        uic.loadUi("inventory.ui", self)
        
        #Define Widget
        self.main = self.findChild(QPushButton, "main")
        self.table = self.findChild(QTableWidget, "tableWidget")
        self.submitInfo = self.findChild(QPushButton, "submitInfo")
        print(UI.data)
        
        
        self.table.setRowCount(16)
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(['Column 1', 'Column 2 (editable)'])
        
        
         # Create the table rows
        for i, (key, value) in enumerate(UI.data.items()):
            for j, cell_data in enumerate([key, value]):
                cell = QTableWidgetItem(str(cell_data))
                if j == 1:
                    print("hello")
                self.table.setItem(i, j, cell)
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 0)
            item.setFlags(item.flags() ^ Qt.ItemIsEditable)
    
        self.table.setColumnWidth(0, 150)
        self.table.setColumnWidth(1, 200)
                        

        
        #Attach action
        self.main.clicked.connect(self.backToMain)
        self.table.cellChanged.connect(self.update_dict)
        self.submitInfo.clicked.connect(self.submitInformation)
        
    def update_list(self):
        self.table.clear()
        # Create the table rows
        for i, (key,value) in enumerate(UI.data.items()):
            for j, cell_data in enumerate([key, value]):
                cell = QTableWidgetItem(str(cell_data))
                if j == 1:
                    print("hello")
                self.table.setItem(i, j, cell)
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 0)
            item.setFlags(item.flags() ^ Qt.ItemIsEditable)
            
    
    
    def update_dict(self):
        for row in range(self.table.rowCount()):
            item1 = self.table.item(row, 0)
            item2 = self.table.item(row, 1)
            if item1 and item2:  # Check if both items exist
                key = item1.text()  # Get the key from the first column
                value = item2.text()  # Get the new value from the second column
                UI.data[key] = value  # Update the dictionary with the new value

    def submitInformation(self):
        # Deselect the current cell to trigger cellChanged signal
        self.table.setCurrentCell(-1, -1)

        # Load the workbook
        workbook = openpyxl.load_workbook('Active Inventory.xlsx')
    
        # Select the worksheet
        worksheet = workbook['All Inventory']
    
        # Remember the currently selected cell
        current_row = self.table.currentRow()
        current_column = self.table.currentColumn()
        
        # Search for a cell containing a specific value
        search_value = UI.data["Tag"]
        valuefound = False
        column = worksheet['D']
        for cell in column:
            if cell.value == search_value:
                coordinate = cell.row
                valuefound = True
        v = list(UI.data.values())
        #update the inventory spreadsheet because already in system
        if(valuefound):
            column_Marker = 1;
            while(column_Marker < 17):
                worksheet.cell(row=coordinate, column=column_Marker).value = v[column_Marker - 1]
                column_Marker = column_Marker + 1

        # Save the workbook
        workbook.save('Active Inventory.xlsx')
    
        # Update the table with the edited data
        self.update_list()
        
        # Restore the selection to the remembered cell
        self.table.setCurrentCell(current_row, current_column)
    
        # Inform the user that their changes have been saved
        QMessageBox.information(self, "Success", "Your changes have been saved.")

    def backToMain(self):
        # Navigate back to the main window
        widget.setCurrentIndex(widget.currentIndex() - 1)



class checkboxSoftware(QDialog):
    def __init__(self):
        super(checkboxSoftware, self).__init__()
        uic.loadUi("checkSoftware.ui", self)
        
        #Define Widget
        self.submit = self.findChild(QPushButton, "submit")
        self.gridbox = self.findChild(QGridLayout, "grid")
        
        self.toCheck = ['sevenzip', 'adobe', 'bomgar', 'citrix', 'code', 'cortex', 'endnote', 'matlab', 'google_chrome', 'microsoft_excel', 'microsoft_onedrive', 'microsoft_powerpoint', 'microsoft_teams', 'microsoft_word', 'mozilla_firefox', 'office365', 'putty', 'qualys', 'zoomus']
        
        self.toCheckUpdate = []
        
        x = 0
        while (x < len(self.toCheck)):
            self.box = self.findChild(QCheckBox, self.toCheck[x])
            if self.box.isChecked():
                self.toCheckUpdate.append(self.box.text())
            x += 1

        
        #Attach action
        self.submit.clicked.connect(lambda: self.checkInstallation(self.toCheckUpdate))


    #Define action
    def checkInstallation(self, toCheckUpdate):
        #run checksoftware and update page
        installed = installedPage(toCheckUpdate)
        widget.addWidget(installed)
        widget.setCurrentIndex(widget.currentIndex() + 2)
        
class installedPage(QDialog):
    def __init__(self, toCheck1):
        super(installedPage, self).__init__()
        uic.loadUi("installedPage.ui", self)
        
        #Define Widget
        self.main = self.findChild(QPushButton, "exitToMain")
        self.installedSoftware = self.findChild(QListWidget, "listWidget")
        self.notInstalledSoftware = self.findChild(QListWidget, "listWidget_2")
        
        list = toCheck1
        def is_app_installed(app_name):
            app_path = f"/Applications/{app_name}.app"
            return os.path.exists(app_path)

        appsInstalled = []
        appsNotInstalled = []
        x = 0
        while (x < len(list)):
            if is_app_installed(list[x]):
                appsInstalled.append(list[x])
            else:
                appsNotInstalled.append(list[x])
            x += 1;

        
        
        # Loop through the list of strings and add them to the list widget
        for string in appsInstalled:
            list_item = QListWidgetItem(string)
            self.installedSoftware.addItem(list_item)

        for string in appsNotInstalled:
            list_item = QListWidgetItem(string)
            self.notInstalledSoftware.addItem(list_item)

        
        #Attach action
        self.main.clicked.connect(self.backToMain4)

    #Define action
    def backToMain4(self):
        widget.setCurrentIndex(widget.currentIndex() - 4)



class InventoryGathering():

    def printRow(row):
        print("\nBuilding: "  + str(row[0] or '')  + "\nRoom: " + str(row[1] or '') + "\nComputer Name: " + str(row[2] or '') + "\nTag: " + str(row[3] or '') + "\nS/N: " + str(row[4] or '') + "\nMAC: " + str(row[5] or '') + "\nwMac: " + str(row[6] or '') + "\nFirst Production Date: " + str(row[7] or '') + "\nMake: " + str(row[8] or '') + "\nModel: " + str(row[9] or '') + "\nLast Service Date: " + str(row[10] or '') + "\nUser: " + str(row[11] or '') + "\nAdvisor: " + str(row[12] or '') + "\nProject: " + str(row[13] or '') + "\nRequistion: " + str(row[14] or '') + "\nNotes: " + str(row[15] or ''))

    def rowList():
        os = platform.system()
        os = os.lower()
        if "win" in os:
            print("Windows")
        elif "lin" in os:
            print("This is a Linux machine")
        else:
            print("This is a Mac")
            InventoryGathering.mac()
    
    def mac():
        #Computer name
        result = subprocess.run(['scutil', '--get', 'ComputerName'], stdout=subprocess.PIPE)
        computer_name = result.stdout.decode().strip()
        computer_name = computer_name.replace('\n', '')
        computer_name = 'ME01W2323ADM01L'
        UI.data["Computer Name"] = computer_name
        
        #parse the computer name
        building = computer_name[2:4]
        os = computer_name[4:5]
        room = computer_name[5:9]
        def switch_case(x):
            print("x = " + x)
            if x == '01':
                return 'MRDC'
            elif x == '02':
                return 'GTMI'
            elif x == '03':
                return 'Love'
            elif x == '04':
                return 'Boggs'
            elif x == '05':
                return 'Bunger Henry'
            elif x == '06':
                return 'SCC'
            elif x == '07':
                return 'GTRI'
            elif x == '08':
                return 'CNES'
            elif x == '09':
                return 'IBB'
            elif x == '10':
                return 'RBI'
            elif x == '11':
                return 'MiRC'
            elif x == '12':
                return 'CCB'
            elif x == '13':
                return 'Whitaker'
            elif x == '14':
                return 'MoS&E'
            elif x == '15':
                return 'TSRB'
            elif x == '16':
                return 'Ben Zinn Building'
            elif x == '17':
                return 'TEP'
            elif x == '18':
                return 'Marcus Nano Building'
            elif x == '19':
                return 'MoSE'
            elif x == '20':
                return 'A.French'
            elif x == '21':
                return 'Rich Computing Center'
            elif x == '22':
                return 'BCDC'
            elif x == '23':
                return 'Ford'
            else:
                return 'No location assigned'

        UI.data["Building"] = switch_case(building)
        if(UI.data["Building"] == ' '):
            UI.data["Room"] = 'No location assigned'
        else:
            UI.data["Room"] = room
            


        # IP address
        result = subprocess.run(['ifconfig'], stdout=subprocess.PIPE)
        output = result.stdout.decode()
        start = output.find("inet ") + 5
        end = output.find(" netmask")
        ip_address = output[start:end]
        print("IP address:", ip_address)
        UI.data["Notes"] = ip_address

        # MAC address
        result = subprocess.run(["ifconfig"], stdout=subprocess.PIPE)
        output = result.stdout.decode()
        start = output.find("ether ") + 6
        end = start + 17
        mac_address = output[start:end]
        mac_address = mac_address.replace(":", "-").upper()
        print("MAC address:", mac_address)
        UI.data["MAC"] = mac_address

        #Serial Number
        # Run the system_profiler command to get the hardware information
        result = subprocess.run(["system_profiler", "SPHardwareDataType"], stdout=subprocess.PIPE)
        # Extract the serial number from the command output
        output = result.stdout.decode()
        start = output.find("Serial Number (system):") + 23
        end = start + 12
        serial_number = output[start:end]
        # Print the serial number
        print("Serial number:", serial_number)
        UI.data["S/N"] = serial_number

        #Model
        # Run the system_profiler command to get the hardware information
        result = subprocess.run(["system_profiler", "SPHardwareDataType"], stdout=subprocess.PIPE)
        # Extract the make and model from the command output
        output = result.stdout.decode()
        start = output.find("Model Name:") + 12
        end = output.find("Model Identifier:") - 1
        make_and_model = output[start:end]
        # Print the make and model
        print("Make and model:", make_and_model)
        UI.data["Model"] = make_and_model

        #Service Date
        today = date.today()
        UI.data["Last Service Date"] = today.strftime("%m/%d/%Y")


        #Make
        UI.data["Make"] = "Apple"

    
    #tag
    def tagger(self, tagInput):
        #tagNumber
        UI.data["Tag"] = tagInput
        InventoryGathering.rowList()
        InventoryGathering.printRow(list(UI.data.values()))
        print(UI.data)
        InventoryGathering.update_inventory_sheet(self, tagInput)
        
    def update_inventory_sheet(self, tagNumber):
        # Load the workbook
        workbook = openpyxl.load_workbook('Active Inventory.xlsx')

        # Select the worksheet
        worksheet = workbook['All Inventory']

        # Search for a cell containing a specific value
        search_value = tagNumber
        valuefound = False
        column = worksheet['D']
        for cell in column:
            if cell.value == search_value:
                coordinate = cell.row
                valuefound = True

        #update the inventory spreadsheet because already in system
        if(valuefound):
            rowValues = []
            column_Marker = 1;
            while(column_Marker < 17):
                cell = worksheet.cell(row=coordinate, column=column_Marker).value
                rowValues.append(cell)
                column_Marker = column_Marker + 1
            
            UI.data["wMac"] = rowValues[6]
            UI.data["First Production Date"] = rowValues[7]
            UI.data["User"] = rowValues[11]
            UI.data["Advisor"] = rowValues[12]
            UI.data["Project"] = rowValues[13]
            UI.data["Requistion"] = rowValues[14]

            # Save the workbook
            workbook.save('Active Inventory.xlsx')
    
        #create new entry
        else:
            #Start Date
            #rowDict["First Production Date"] = today.strftime("%m/%d/%Y")


            # Define a list of new values
            new_values = list(UI.data.values())
    
            worksheet.append(new_values)
    
            # Save the workbook
            workbook.save('Active Inventory.xlsx')




# Initialize The App
app = QApplication(sys.argv)
widget = QtWidgets.QStackedWidget()
UIWindow = UI()
inventory = Inventory()
tag = Tag()
checkbox = checkboxSoftware()

widget.addWidget(UIWindow)
widget.addWidget(inventory)
widget.addWidget(checkbox)
widget.addWidget(tag)

#Widget stack
#0 - main
#1 - inventory
#2- checkbox
#3 - tag
#4 - installed page


widget.setFixedHeight(330)
widget.setFixedWidth(514)
widget.show()
app.exec_()
